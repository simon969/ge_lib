import io
import pandas as pd
import math 
import json
from rich import print as rprint


tabled_groups = {'groups': ['core', 'dcpt', 'dcpg','eres','gchm','ispt', 'lden', 'llpl', 'lnmc', 'lvan', 'mond', 'rden', 'trit', 'wstd'],
                'formats': ['csv','json','xml','xlsx']
                }

def get_data_table (input_file, tables, table, errors):
    
    if table == 'core':
        return data_core (input_file, tables, errors)   
    if table == 'dcpt':
        return data_dcpt (input_file, tables, errors)  
    if table == 'dcpg':
        return data_dcpg (input_file, tables, errors) 
    if table == 'eres':
        return data_eres (input_file, tables, errors)  
    if table == 'gchm':
        return data_gchm (input_file, tables, errors)   
    if table == 'ispt':
        return data_ispt (input_file, tables, errors)
    if table == 'lden':
        return data_lden (input_file, tables, errors)   
    if table == 'llpl':
        return data_llpl (input_file, tables, errors)   
    if table == 'lnmc':
        return data_lnmc (input_file, tables, errors)
    if table == 'lvan':
        return data_lvan (input_file, tables, errors)
    if table == 'mond':
        return data_mond (input_file, tables, errors)
    if table == 'rden':
        return data_rden (input_file, tables, errors)  
    if table == 'trit':
        return data_trit (input_file, tables, errors)
    if table == 'wstd':
        return data_wstd (input_file, tables, errors)

def str_to_bytes(str:str):
    return bytes (str, 'utf-8')
def json_to_bytes(dict:dict):
    return bytes (json.dumps(dict), 'utf-8')

def dict_to_xml (dict:list):
    
    # https://pypi.org/project/dict2xml/
    from dict2xml import dict2xml

    xml = dict2xml(data=dict, 
                   wrap ='row', 
                   indent ='   ')
    prolog = '<?xml version=\"1.0\" encoding=\"UTF-8\"?>'

    return prolog + '<root>' + xml + '</root>'

def dict_to_xlsx(dict:dict):
    
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    io_stream = io.BytesIO()
    
    # write header
    for column, value in enumerate(dict[0].keys()):
            ws.cell(row=1,column=column + 1).value = value

    # write values
    for row, item in enumerate(dict):
        for column, (key, value) in enumerate(item.items()):
            ws.cell(row=row + 2, column=column + 1).value = value
    
    wb.save(io_stream)
    
    io_stream.seek(0)
    return io_stream.getvalue()

def dict_to_csv(dict:dict):

    # https://stackoverflow.com/questions/10373247/how-do-i-write-a-python-dictionary-to-a-csv-file
    
    from csv import DictWriter, QUOTE_NONNUMERIC

    file_out = io.StringIO()
    w = DictWriter (f = file_out, 
                       fieldnames = dict[0].keys(),
                       quoting=QUOTE_NONNUMERIC)
    w.writeheader()
    w.writerows(dict)
    file_out.seek(0)
    
    return file_out.getvalue()



def get_group(tables, table, errors):
    try :
        grp = tables[table]
        return grp
    except:    
        error = {"description":"table {} not found".format(table)}
        errors.append (error)
        return None
def get_group_data(tables, table, where, errors):
    grp = get_group(tables, table, errors)
    if grp is not None:
        qry = grp.query(where)
        return qry
    else:
        return None
def get_value(table, where, field, value_not_found = None):
    try:
        rows =  table.query(where) 
        if len(rows.index) > 0:
            val = pd.to_numeric(rows[field][0], errors='raise')
            if val == 'NaN':
                return value_not_found
            return val
        else:
            return value_not_found
    except:
        return value_not_found
def get_str(table, where, field, str_not_found = None):
    try:
        rows =  table.query(where) 
        if len(rows.index) > 0:
            val = rows[field][0]
            return val
        else:
            return str_not_found
    except:
        return str_not_found
def get_str_list (table, fields, str_not_found = None):
    list = []
    for f in fields:
        try:
            list.append (table[f])
        except:
            list.append(str_not_found)
    return list

def get_str_dict (table, fields, str_not_found = None):
    dict = {}
    for f in fields:
        try:
            dict[f] = table[f]
        except:
            dict[f] = str_not_found
    return dict
def get_headers_dict (table, fields, not_found=None):
    dict = {}
    for f in fields:
        try:
            col = table[f]
            dict[f] = {"unit":col[1],
                        "type":col[2]}
        except:
            dict[f] =  {"unit":not_found,
                        "type":not_found}
    return dict

def get_value_list (table, fields, not_float = math.nan):    
    
    def try_float(var, not_float):
        try:
            if var is None:
                return not_float
            return float(var)
        except ValueError:
            return not_float

    str_list = get_str_list(table, fields, None)
    
    float_list = [try_float(i, not_float) for i in str_list]
    
    return float_list


def add_coords (df):
    df1 =  df.copy()
    
    df1['loc_x'] = df[['LOCA_LOCX','LOCA_NATE']].max(axis=1)
    df1['loc_y'] = df[['LOCA_LOCY','LOCA_NATN']].max(axis=1)
    df1['loc_z'] = df[['LOCA_LOCZ','LOCA_GL']].max(axis=1)
    
    return df1

def get_coords (loca):
    
    values_z = get_value_list (loca,['LOCA_LOCZ','LOCA_GL'],None)
    values_y =  get_value_list (loca,['LOCA_LOCY','LOCA_NATN'],None)
    values_x =  get_value_list (loca,['LOCA_LOCX','LOCA_NATE'],None)

    return max(values_x), max(values_y), max(values_z)

def get_group_by (df, value_fields, group_fields):
    list = []
    for v in value_fields:
        subdf = df.loc[df[v] != ""]
        for g in group_fields:
            subdf2 =  subdf.loc[subdf[g] != ""]
            if len(subdf2.index) > 0:
                if g not in list:
                    list.append (g)
    return list

def is_numeric(x):
    return isinstance(x, (int, float, complex))

    
def ground_level (loca, loca_fields, none_value):

    if 'LOCA_LOCZ' in loca and 'LOCA_LOCZ' in loca_fields:
        loca_locz = pd.to_numeric(loca['LOCA_LOCZ'])
        if not pd.isnull(loca_locz) and is_numeric (loca_locz):
            return loca_locz 

    if 'LOCA_GL' in loca and 'LOCA_GL' in loca_fields:
        loca_gl = pd.to_numeric(loca['LOCA_GL'])
        if not pd.isnull(loca_gl) and is_numeric (loca_gl):
            return loca_gl
        
    return none_value

def calc_depth_dcpg (values, depth_fields):
    pen = pd.to_numeric(values['DCPT_PEN'])
    test_depth = pd.to_numeric(values['DCPG_DPTH'])
    depth = test_depth + pen/1000
    return depth

def calc_depth_wstd (values, depth_fields):
    depth = pd.to_numeric(values['WSTD_POST'])
    return depth

def calc_depth_default (values, depth_fields):
    depth_res = get_value_list(values, depth_fields)
    depth = max ([float(i) for i in depth_res])
    return depth

def calc_level_default (depth, loca, loca_fields, values):
    gl = ground_level(loca.iloc[0], loca_fields,0.0)
    level = float(gl)-float(depth)
    return level

def get_df( tables, 
            table, 
            value_fields=[],
            depth_fields=['SAMP_TOP','SPEC_DPTH'],
            calc_depth = calc_depth_default,
            calc_level = calc_level_default,
            loca_fields = ['LOCA_LOCX','LOCA_LOCY','LOCA_LOCZ','LOCA_NATE','LOCA_NATN','LOCA_GL'],
            geol_fields = ['GEOL_GEOL','GEOL_GEO2','GEOL_GEO3','GEOL_LEG'], 
            nan_value = 0.0,
            errors=[]) :
    '''
     Parameters
    ----------
    tables: dataframe tables 
    table: primary ags group to return
    value_fields: values fields from primary ags group
    depth_fields: fields in primary group table that should be used to find associated records in grology table 
                  (if more than one field provided the deepest will be used)
    calc_depth: function to be used to calculate depth
    calc_level:function to be used to calculate level
    loca_fields: fields to be brought from loca group
    geol_fields: fields to be brought from geol group
    nan_value: value that should be used for values that are not numbers
    errors[]: any errors encountered will be added to this array
    
    Returns
    -------
    get_def returns a combined dataframe of the primary table and related record values from the geology and loca groups
    
    
    '''
    import numpy as np
    try:

        grp = get_group(tables, table, errors)
        loc = get_group_data(tables, 'LOCA', "HEADING == 'DATA'", errors)
        geo = get_group_data(tables, 'GEOL', "HEADING == 'DATA'", errors)
        
        if grp is None:
            errors.append ("table group {0} not found in groups{1}".format(table, tables.keys()))
            return None
        
        else:
            grp = grp.copy()

        if geo is not None:
            geo = geo.copy()
            geo.loc[:,'GEOL_TOP'] = geo.loc[:,'GEOL_TOP'].astype(float)
            geo.loc[:,('GEOL_BASE')] = geo.loc[:,('GEOL_BASE')].astype(float)
        
        if loc is not None:
            loc = loc.copy()
        
        for f in geol_fields + loca_fields + ['level','depth']:
            if f not in grp.columns:
                grp[f] = ''
                
        for index, values in grp.query ("HEADING == 'DATA'").iterrows():
            
            location = values['LOCA_ID']
            level = ''

            geol_res = []
            loca_res = []

            depth = calc_depth (values, depth_fields)
            grp.at[index,'depth'] = depth
            
            if loc is not None:
                where = "LOCA_ID=='{0}'".format(values['LOCA_ID'])
                loca = loc.query(where)
                if len(loca.index)>0:
                    level = calc_level (depth, loca, loca_fields, values)
                    grp.at[index,'level'] = level
                    loca_res = get_value_list (loca.iloc[0],loca_fields,None)
                for field, val in zip (loca_fields,loca_res):
                    grp.at[index,field] = val
            
            if geo is not None:
                where = "LOCA_ID=='{0}' and GEOL_TOP<={1} and GEOL_BASE>={2}".format(location,depth,depth)
                geo_qry = geo.query(where)
                if len(geo_qry.index)>0:
                    geol_res = get_str_list (geo_qry.iloc[0], geol_fields)
                for field, val in zip(geol_fields,geol_res):
                    grp.at[index,field] = val
            
        for f in value_fields: 
            for index, values in grp.query ("HEADING == 'DATA'").iterrows():
                grp.at[index,f] = pd.to_numeric(values[f], errors='coerce')
        
        grp.query("HEADING == 'DATA'").fillna(nan_value)
        
        return grp
    
    except Exception as e:
        s = str(e)
        errors.append(s)

def get_data (input_file=[],
              tables=None,
              table=None, 
              value_fields=[],
              other_fields=[],
              depth_fields=['SAMP_TOP','SPEC_DPH'],
              loca_fields=['LOCA_LOCX','LOCA_LOCY','LOCA_LOCZ'],
              geol_fields=['GEOL_GEOL','GEOL_GEO2','GEOL_GEO3','GEOL_LEG'],
              errors =[]):
    '''
    Parameters
    ----------
    tables: dataframe tables if they have already been read from the input_file. 
            If the input file is an array this must also to be an array or None and each file will be read
    table: primary ags group to return
    value_fields: value fields to be returned from the primary table
    
    Returns
    -------
    returns a combined list of the primary table and related record values from the geology and loca groups
    '''

    import pandas as pd
    import os
    from python_ags4.AGS4 import AGS4_to_dataframe

    if _is_array_like(input_file):
        files = input_file
    else:
        files = [input_file]
    
    for file in files:

        if tables is None:
            tables, headers = AGS4_to_dataframe (file, get_line_numbers=False)
        
        results = []

        df = get_df (tables,
                table, 
                value_fields=[], 
                depth_fields=depth_fields,
                loca_fields=loca_fields,
                geol_fields=geol_fields,
                errors=errors)
        
        if df is None:
            errors.append ("get_df returned None for group {0}".format(table))
            return results
        
        if _is_file_like(input_file):
            file_name = os.path.basename(input_file.name)
        else:
            file_name = input_file

        for index, values in df.query ("HEADING == 'DATA'").iterrows():
                res = {'file':file_name,
                            'location':values['LOCA_ID'],
                            'table':table,
                            'level':values['level'],
                            'depth':values['depth']
                            }
                value_res = get_value_list(values, value_fields, None)
                for field, val in zip (value_fields,value_res): 
                    res[field] = val
                other_res = get_str_list(values, other_fields, None)
                for field, val in zip (other_fields,other_res): 
                    res[field] = val
                depth_res = get_value_list(values, depth_fields, None)
                for field, val in zip (depth_fields,depth_res):
                    res[field] = val
                geol_res = get_str_list (values, geol_fields)
                for field, val in zip(geol_fields,geol_res):
                    res[field] = val            
                loca_res = get_str_list (values, loca_fields)
                for field, val in zip (loca_fields,loca_res):
                    res[field] = val
                
                results.append(res)

    return results

def get_all_data (input_file=[],
              tables = None,
              table = None, 
              value_fields=[],
              other_fields=[],
              depth_fields=['SAMP_TOP','SPEC_DPH'],
              loca_fields=['LOCA_LOCX','LOCA_LOCY','LOCA_LOCZ'],
              geol_fields=['GEOL_GEOL','GEOL_GEO2','GEOL_GEO3','GEOL_LEG'],
              errors =[]):
    ''' 
    parameters
    ----------
    input_files:
    string or open buffer list of ags files to read table

    tables:
    dataframe tables
    table:
    returns
    -------

    list of json strings

    '''
    import pandas as pd
    import os
    from python_ags4.AGS4 import AGS4_to_dataframe
    
    results = []

    if _is_array_like(input_file):
        files = input_file 
    else:
        files = [input_file]
    
    for file in files:

        if tables is None:
            tables, headers = AGS4_to_dataframe (file, get_line_numbers=False)

        df = get_df (tables,
                table, 
                value_fields=[], 
                depth_fields=depth_fields,
                loca_fields=loca_fields,
                geol_fields=geol_fields,
                errors=errors)
        if df:
            if _is_file_like(file):
                file_name = os.path.basename(file.name)
            else:
                file_name = input_file

            for index, values in df.query ("HEADING == 'DATA'").iterrows():
                    res = {"file":file_name,
                                "table":table
                                }
                    for field, val in zip (df.columns,values): 
                        res[field] = val
                    
                    results.append(res)

    return results
def _is_array_like (obj):
    """Check if object is file like
    
    returns
    -------
    bool
        Return True if obj is file like, otherwise return False
    """
    if obj.__class__ == list:
        return True
    

    return False
def _is_file_like(obj):
    """Check if object is file like

    returns
    -------
    bool
        Return True if obj is file like, otherwise return False
    """

    if not (hasattr(obj, 'read') or hasattr(obj, 'write')):
        return False

    if not hasattr(obj, "__iter__"):
        return False

    return True

def data_ispt(input_file, tables, errors):
   
    rprint('[green] Getting data from ISPT...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'ISPT',
                     value_fields = ['ISPT_NVAL'],
                     depth_fields = ['ISPT_TOP'], 
                     errors = errors)
def data_trit(input_file, tables, errors):
   
    rprint('[green] Getting data from TRIT...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'TRIT',
                     value_fields = ['TRIT_CU','TRIT_BDEN','TRIT_BDEN'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)
def data_core(input_file, tables, errors):
   
    rprint('[green] Getting data from CORE...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'CORE',
                     value_fields= ['CORE_PREC','CORE_SREC','CORE_RQD'],
                     depth_fields = ['CORE_TOP','CORE_BASE'], 
                     errors = errors)

def data_dcpt(input_file, tables, errors):
   
    rprint('[green] Getting data from DCPT...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'DCPT',
                     value_fields = ['DCPT_CBLO'],
                     depth_fields = ['DCPG_DPTH'], 
                     errors = errors)
def data_dcpg(input_file, tables, errors):
   
    rprint('[green] Getting data from DCPG...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'DCPG',
                     value_fields = ['DCPG_ZERO'],
                     other_fields = ['DCPG_DATE','DCPG_TESN'],
                     depth_fields = ['DCPG_DPTH'], 
                     errors = errors)
def data_eres(input_file, tables, errors):
   
    rprint('[green] Getting data from ERES...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'ERES',
                     value_fields = ['ERES_RVAL'],
                     other_fields = ['ERES_CODE','ERES_NAME','ERES_RUNI','ERES_RTXT','ERES_DTIM'],
                     errors = errors)
def data_gchm(input_file, tables, errors):
   
    rprint('[green] Getting data from GCHM...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'GCHM',
                     value_fields = ['ERES_RVAL'],
                     other_fields= ['GCHM_CODE','GCHM_METH','GCHM_TTYP','GCHM_RESL','GCHM_UNIT','GCHM_NAME'],
                     errors = errors)
def data_lden(input_file, tables, errors):
   
    rprint('[green] Getting data from LDEN...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'LDEN',
                     value_fields = ['LDEN_BDEN'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)

def data_llpl(input_file, tables, errors):
   
    rprint('[green] Getting data from LLPL...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'LLPL',
                     value_fields = ['LLPL_LL','LLPL_PL','LLPL_PI'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)
def data_lnmc(input_file, tables, errors):
   
    rprint('[green] Getting data from LNMC...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'LNMC',
                     value_fields = ['LNMC_NMC'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)
def data_lvan(input_file, tables, errors):
   
    rprint('[green] Getting data from LVAN...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = "LVAN",
                     value_fields = ['LVAN_VNPK','LVAN_VNRM','LVAN_MC'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)
def data_mond(input_file, tables, errors):
   
    rprint('[green] Getting data from MOND...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'MOND',
                     value_fields = ['MOND_DTIM','MOND_TYPE','MOND_RDNG','MOND_UNIT'],
                     depth_fields = ['MONG_DIS'], 
                     errors = errors)
def data_rden(input_file, tables, errors):
   
    rprint('[green] Getting data from RDEN...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'RDEN',
                     value_fields = ['RDEN_MC','RDEN_SMC','RDEN_BDEN','RDEN_DDEN','RDEN_PORO','RDEN_PDEN'],
                     depth_fields = ['SAMP_TOP','SPEC_DPTH'], 
                     errors = errors)

def data_wstd(input_file, tables, errors):
   
    rprint('[green] Getting data from WSTD...[/green]')
    return get_data (input_file=input_file,
                     tables=tables, 
                     table = 'WSTD',
                     value_fields = ['WSTD_NMIN','WSTD_POST'],
                     depth_fields = ['WSTG_DPTH'], 
                     errors = errors)    